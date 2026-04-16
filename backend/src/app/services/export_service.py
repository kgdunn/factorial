"""Render an experiment to PDF / XLSX / CSV / Markdown byte streams.

Callers:
- ``backend/src/app/api/v1/endpoints/experiments.py`` (owner exports)
- ``backend/src/app/api/v1/endpoints/shares_public.py`` (public exports)

The PDF and the HTML snapshot on the public share page are rendered
from the *same* Jinja templates under ``backend/templates/`` so they
stay visually identical.
"""

from __future__ import annotations

import base64
import csv
import io
import logging
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status
from jinja2 import Environment, FileSystemLoader, select_autoescape

from app.config import settings
from app.models.experiment import Experiment
from app.services.chart_render_service import render_echarts_png

logger = logging.getLogger(__name__)

_TEMPLATES_DIR = Path(__file__).resolve().parents[3] / "templates"

_jinja_env = Environment(
    loader=FileSystemLoader(str(_TEMPLATES_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
)


# ---------------------------------------------------------------------------
# Shared helpers — build the "report context" both PDF and HTML consume
# ---------------------------------------------------------------------------


def _design_rows(experiment: Experiment, *, include_results: bool) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (column_names, rows) merging design matrix with results.

    ``design_actual`` wins over ``design_coded`` when present.  Results
    are merged by ``run_index``.
    """
    design_data = experiment.design_data or {}
    matrix = design_data.get("design_actual") or design_data.get("design_coded") or []
    results_list = experiment.results_data or []
    results_by_idx = {r.get("run_index"): r for r in results_list}

    factor_cols: list[str] = []
    response_cols: list[str] = []
    for row in matrix:
        for k in row:
            if k not in factor_cols:
                factor_cols.append(k)

    if include_results:
        for r in results_list:
            for k in r:
                if k != "run_index" and k not in factor_cols and k not in response_cols:
                    response_cols.append(k)

    columns = ["run_index", *factor_cols, *response_cols]
    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(matrix):
        merged: dict[str, Any] = {"run_index": idx, **row}
        if include_results:
            merged.update({k: v for k, v in results_by_idx.get(idx, {}).items() if k != "run_index"})
        rows.append(merged)
    return columns, rows


def _iter_echarts_options(experiment: Experiment) -> list[dict[str, Any]]:
    """Best-effort extraction of ECharts option dicts from design_data.

    We do not require any particular schema — anything shaped like
    ``{"echarts": {...}}`` or a top-level ``plots`` array is accepted.
    """
    design_data = experiment.design_data or {}
    options: list[dict[str, Any]] = []

    plots = design_data.get("plots")
    if isinstance(plots, list):
        for p in plots:
            if isinstance(p, dict) and isinstance(p.get("echarts"), dict):
                options.append(p["echarts"])
            elif isinstance(p, dict) and "xAxis" in p:
                # raw option dict stored directly
                options.append(p)

    if not options and isinstance(design_data.get("echarts"), dict):
        options.append(design_data["echarts"])

    return options


async def _render_plots_as_data_uris(experiment: Experiment) -> list[dict[str, str]]:
    """Pre-render each ECharts option to a PNG and return data URIs + caption."""
    options = _iter_echarts_options(experiment)
    rendered: list[dict[str, str]] = []
    for idx, option in enumerate(options):
        try:
            png = await render_echarts_png(option)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Plot %d render failed: %s", idx, exc)
            continue
        b64 = base64.b64encode(png).decode("ascii")
        # Playwright returns real PNG; the fallback is SVG text.
        mime = "image/png" if png.startswith(b"\x89PNG") else "image/svg+xml"
        title_dict = option.get("title") if isinstance(option.get("title"), dict) else None
        title = title_dict.get("text", f"Plot {idx + 1}") if title_dict else f"Plot {idx + 1}"
        rendered.append(
            {
                "index": str(idx),
                "data_uri": f"data:{mime};base64,{b64}",
                "title": title,
            }
        )
    return rendered


def _report_context(
    experiment: Experiment,
    *,
    include_results: bool,
    share_url: str | None,
    plots: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    columns, rows = _design_rows(experiment, include_results=include_results)
    design_data = experiment.design_data or {}
    return {
        "experiment": {
            "id": str(experiment.id),
            "name": experiment.name,
            "status": experiment.status,
            "design_type": experiment.design_type,
            "n_factors": design_data.get("n_factors"),
            "n_runs": design_data.get("n_runs"),
            "factors": experiment.factors or [],
            "created_at": experiment.created_at,
        },
        "columns": columns,
        "rows": rows,
        "include_results": include_results,
        "share_url": share_url,
        "plots": plots or [],
    }


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def build_csv(experiment: Experiment, *, include_results: bool = True) -> bytes:
    columns, rows = _design_rows(experiment, include_results=include_results)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def build_xlsx(experiment: Experiment, *, include_results: bool = True) -> bytes:
    from openpyxl import Workbook  # noqa: PLC0415

    wb = Workbook()

    design_sheet = wb.active
    design_sheet.title = "Design"
    columns, rows = _design_rows(experiment, include_results=False)
    design_sheet.append(columns)
    for row in rows:
        design_sheet.append([row.get(c) for c in columns])

    if include_results:
        responses_sheet = wb.create_sheet("Responses")
        rcols, rrows = _design_rows(experiment, include_results=True)
        responses_sheet.append(rcols)
        for row in rrows:
            responses_sheet.append([row.get(c) for c in rcols])

    meta_sheet = wb.create_sheet("Metadata")
    meta_sheet.append(["Field", "Value"])
    meta_sheet.append(["name", experiment.name])
    meta_sheet.append(["design_type", experiment.design_type or ""])
    meta_sheet.append(["status", experiment.status])
    design_data = experiment.design_data or {}
    meta_sheet.append(["n_factors", design_data.get("n_factors", "")])
    meta_sheet.append(["n_runs", design_data.get("n_runs", "")])
    meta_sheet.append(["created_at", experiment.created_at.isoformat() if experiment.created_at else ""])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------


def build_markdown(
    experiment: Experiment,
    *,
    include_results: bool = True,
    share_url: str | None = None,
) -> bytes:
    template = _jinja_env.get_template("experiment_report.md.j2")
    ctx = _report_context(
        experiment,
        include_results=include_results,
        share_url=share_url,
        plots=None,
    )
    return template.render(**ctx).encode("utf-8")


# ---------------------------------------------------------------------------
# HTML (shared by PDF and the public share page)
# ---------------------------------------------------------------------------


async def build_html(
    experiment: Experiment,
    *,
    include_results: bool = True,
    share_url: str | None = None,
    render_plots: bool = True,
) -> str:
    template = _jinja_env.get_template("experiment_report.html.j2")
    plots = await _render_plots_as_data_uris(experiment) if render_plots else []
    ctx = _report_context(
        experiment,
        include_results=include_results,
        share_url=share_url,
        plots=plots,
    )
    return template.render(**ctx)


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


async def build_pdf(
    experiment: Experiment,
    *,
    include_results: bool = True,
    share_url: str | None = None,
) -> bytes:
    if not settings.exports_enable_pdf:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="PDF export disabled on this deployment",
        )
    try:
        from weasyprint import HTML  # noqa: PLC0415
    except Exception as exc:  # noqa: BLE001
        logger.warning("WeasyPrint import failed: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="PDF export unavailable: WeasyPrint not installed",
        ) from exc

    html = await build_html(
        experiment,
        include_results=include_results,
        share_url=share_url,
        render_plots=True,
    )
    return HTML(string=html).write_pdf()
