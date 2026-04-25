"""Parse user-supplied .xlsx and .csv files into a 2D matrix.

This is a deliberately small pure-function module. It does not call
Claude, does not touch the database, and does not know what factors,
responses, or experiments are. Its only job is to turn raw bytes into
a rectangular ``list[list[Any]]`` that the next layer
(:mod:`app.services.upload_claude_service`) can hand to Claude.

The module also enforces the cell budget that bounds the cost of every
Claude call: refusing oversized files here means we never burn tokens
on a runaway sheet.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook

# Import settings lazily inside the function bodies so this module remains
# trivially importable in unit tests without a populated environment.


SUPPORTED_EXTENSIONS = (".xlsx", ".csv")


class UploadValidationError(ValueError):
    """User-facing error from ``parse_upload``.

    The message is intended to be safe to surface verbatim in the UI.
    """


@dataclass(frozen=True)
class ParsedUpload:
    """Result of :func:`parse_upload`."""

    rows: list[list[Any]]
    detected_format: str


def parse_upload(
    filename: str,
    content: bytes,
    *,
    max_bytes: int | None = None,
    max_cells: int | None = None,
) -> ParsedUpload:
    """Decode ``content`` into a rectangular 2D matrix.

    ``max_bytes`` and ``max_cells`` default to the values on
    :class:`app.config.Settings` (``upload_max_bytes`` and
    ``upload_max_cells``). They are arguments so unit tests can pin
    them without touching the global settings object.

    Raises :class:`UploadValidationError` for any user-actionable
    failure (oversized, unsupported extension, malformed, non-rectangular).
    """

    if max_bytes is None or max_cells is None:
        from app.config import settings

        if max_bytes is None:
            max_bytes = settings.upload_max_bytes
        if max_cells is None:
            max_cells = settings.upload_max_cells

    if len(content) > max_bytes:
        raise UploadValidationError(  # noqa: TRY003
            f"File is {len(content):,} bytes; the limit is {max_bytes:,}."
        )

    lower = (filename or "").lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        raise UploadValidationError(  # noqa: TRY003
            "Legacy .xls files are not supported. Open the file in Excel and save as .xlsx, then upload again."
        )
    if lower.endswith(".xlsx"):
        rows = _parse_xlsx(content)
        detected_format = "xlsx"
    elif lower.endswith(".csv"):
        rows = _parse_csv(content)
        detected_format = "csv"
    else:
        raise UploadValidationError(  # noqa: TRY003
            f"Unsupported file extension. Expected one of {', '.join(SUPPORTED_EXTENSIONS)}."
        )

    rows = _normalise(rows)

    if not rows:
        raise UploadValidationError("The file appears to be empty.")  # noqa: TRY003

    n_cells = sum(len(r) for r in rows)
    if n_cells > max_cells:
        raise UploadValidationError(  # noqa: TRY003
            f"File has {n_cells:,} cells; the limit is {max_cells:,}. Trim the sheet and try again."
        )

    width = len(rows[0])
    for r in rows:
        if len(r) != width:
            raise UploadValidationError(  # noqa: TRY003
                "Rows have different lengths. Make sure every row has the same number of columns."
            )

    return ParsedUpload(rows=rows, detected_format=detected_format)


# ---------------------------------------------------------------------------
# .xlsx
# ---------------------------------------------------------------------------


def _parse_xlsx(content: bytes) -> list[list[Any]]:
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises a wide hierarchy
        raise UploadValidationError(  # noqa: TRY003
            "Could not read the Excel file. It may be corrupt or password-protected."
        ) from exc

    try:
        sheet = _first_non_empty_sheet(wb)
        if sheet is None:
            raise UploadValidationError("The workbook has no non-empty sheets.")  # noqa: TRY003
        return [[_coerce_cell(c) for c in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        wb.close()


def _first_non_empty_sheet(wb):
    for name in wb.sheetnames:
        sheet = wb[name]
        if sheet.max_row and sheet.max_column:
            return sheet
    return None


def _coerce_cell(value: Any) -> Any:
    """Return a JSON-safe scalar for ``value``.

    openpyxl yields ``datetime``, ``date``, ``time``, ``timedelta`` and
    ``Decimal`` instances that ``json.dumps`` cannot serialise. The
    rest of the pipeline only ever shows these to Claude or echoes
    them back to the UI, so loss of precision is fine.
    """

    if value is None:
        return None
    if isinstance(value, dt.datetime | dt.date | dt.time):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    if isinstance(value, Decimal):
        return float(value)
    return value


# ---------------------------------------------------------------------------
# .csv
# ---------------------------------------------------------------------------


def _parse_csv(content: bytes) -> list[list[Any]]:
    text = _decode_csv(content)
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def _decode_csv(content: bytes) -> str:
    """Decode bytes as UTF-8 (with BOM tolerance), falling back to latin-1."""

    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            return content.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise UploadValidationError(  # noqa: TRY003
                "Could not decode the CSV file. Save it as UTF-8 and try again."
            ) from exc


# ---------------------------------------------------------------------------
# Shared post-processing
# ---------------------------------------------------------------------------


def _normalise(rows: list[list[Any]]) -> list[list[Any]]:
    """Strip trailing wholly-empty rows and right-pad short rows.

    openpyxl's used-range can include trailing blanks on either axis
    when a sheet had transient content; a CSV with ragged rows can
    similarly produce uneven widths. We pad to the widest row and
    drop rows that are entirely empty afterwards.
    """

    cleaned = [[_blank_to_none(c) for c in row] for row in rows]
    if not cleaned:
        return []
    width = max((len(r) for r in cleaned), default=0)
    padded = [r + [None] * (width - len(r)) for r in cleaned]
    while padded and all(c is None or c == "" for c in padded[-1]):
        padded.pop()
    return padded


def _blank_to_none(value: Any) -> Any:
    if isinstance(value, str) and value.strip() == "":
        return None
    return value
