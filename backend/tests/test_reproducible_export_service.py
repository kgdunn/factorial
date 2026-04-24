"""Unit tests for ``app.services.reproducible_export_service``.

The fixture stands up an in-memory SQLite, seeds a Conversation +
Experiment + a mix of analysis and non-analysis ``ToolCall`` rows, and
exercises:

- the ``fetch_analysis_tool_calls`` filter/ordering query;
- the rendered ``.py`` source (header metadata + body compiles);
- a round-trip: the rendered script, when exec'd against a stubbed
  ``process_improve.tool_spec.execute_tool_call``, dispatches exactly
  the stored ``tool_input`` in the stored order.
"""

from __future__ import annotations

import io
import sys
import types
import uuid
from datetime import UTC, datetime

import pytest
from fastapi import HTTPException
from sqlalchemy import ColumnDefault
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.db.base import Base
from app.models.conversation import Conversation, ToolCall
from app.models.experiment import Experiment
from app.models.user import User
from app.services import reproducible_export_service


@pytest.fixture
async def db_session():
    engine = create_async_engine(
        "sqlite+aiosqlite:///:memory:",
        connect_args={"check_same_thread": False},
    )

    # Replace PG-side ``gen_random_uuid()`` defaults with a Python-side
    # ``uuid.uuid4`` so SQLite can fill PK columns without the extension.
    for table in Base.metadata.tables.values():
        for col in table.columns:
            if col.server_default is not None and "gen_random_uuid" in str(getattr(col.server_default, "arg", "")):
                col.server_default = None
                col.default = ColumnDefault(uuid.uuid4)

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with factory() as session:
        yield session
    await engine.dispose()


async def _seed(
    session: AsyncSession,
    *,
    tool_calls: list[dict],
    with_conversation: bool = True,
) -> Experiment:
    """Insert a user, conversation, experiment, and tool calls; return the experiment."""
    user = User(
        id=uuid.uuid4(),
        email=f"user-{uuid.uuid4().hex[:6]}@example.com",
        password_hash="x",  # noqa: S106 — fixture value, not a real credential
        display_name="Test User",
    )
    session.add(user)
    await session.flush()

    conversation_id: uuid.UUID | None = None
    if with_conversation:
        conv = Conversation(id=uuid.uuid4(), user_id=user.id, title="conv")
        session.add(conv)
        await session.flush()
        conversation_id = conv.id

    exp = Experiment(
        id=uuid.uuid4(),
        user_id=user.id,
        name="My Experiment",
        status="active",
        design_type="full_factorial",
        factors=[],
        design_data={"n_factors": 2, "n_runs": 4},
        results_data=[],
        conversation_id=conversation_id,
    )
    session.add(exp)
    await session.flush()

    for rec in tool_calls:
        assert conversation_id is not None or rec.get("conversation_id") is not None
        session.add(
            ToolCall(
                id=uuid.uuid4(),
                conversation_id=rec.get("conversation_id") or conversation_id,
                tool_name=rec["tool_name"],
                tool_input=rec.get("tool_input"),
                tool_output=rec.get("tool_output"),
                status=rec.get("status", "success"),
                agent_turn=rec["agent_turn"],
                call_order=rec["call_order"],
            )
        )
    await session.flush()
    return exp


async def test_fetch_analysis_tool_calls_filters_and_orders(db_session):
    # Mix analysis + non-analysis tools across turns, deliberately out of
    # insertion order; the query must filter non-analysis out and sort
    # by (agent_turn, call_order).
    exp = await _seed(
        db_session,
        tool_calls=[
            {"tool_name": "simulate_process", "agent_turn": 1, "call_order": 1, "tool_input": {}},
            {"tool_name": "generate_design", "agent_turn": 1, "call_order": 2, "tool_input": {"a": 1}},
            {"tool_name": "evaluate_design", "agent_turn": 3, "call_order": 1, "tool_input": {"b": 2}},
            {"tool_name": "create_simulator", "agent_turn": 2, "call_order": 1, "tool_input": {}},
            {"tool_name": "analyze_results", "agent_turn": 2, "call_order": 2, "tool_input": {"c": 3}},
            # Failed analysis call — must be dropped.
            {
                "tool_name": "evaluate_design",
                "agent_turn": 4,
                "call_order": 1,
                "tool_input": {"d": 4},
                "status": "error",
            },
        ],
    )

    calls = await reproducible_export_service.fetch_analysis_tool_calls(db_session, exp)

    assert [(c.tool_name, c.agent_turn, c.call_order) for c in calls] == [
        ("generate_design", 1, 2),
        ("analyze_results", 2, 2),
        ("evaluate_design", 3, 1),
    ]


async def test_fetch_returns_empty_when_no_conversation(db_session):
    exp = await _seed(db_session, tool_calls=[], with_conversation=False)
    calls = await reproducible_export_service.fetch_analysis_tool_calls(db_session, exp)
    assert calls == []


async def test_build_python_script_raises_400_without_calls(db_session):
    exp = await _seed(db_session, tool_calls=[])
    with pytest.raises(HTTPException) as info:
        await reproducible_export_service.build_python_script(db_session, exp)
    assert info.value.status_code == 400


async def test_rendered_script_has_header_and_compiles(db_session):
    exp = await _seed(
        db_session,
        tool_calls=[
            {
                "tool_name": "generate_design",
                "agent_turn": 1,
                "call_order": 1,
                "tool_input": {"design_type": "full_factorial", "n_factors": 2},
            },
        ],
    )
    payload = await reproducible_export_service.build_python_script(db_session, exp)
    src = payload.decode("utf-8")

    # Header metadata present.
    assert "Reproducible analysis for experiment: My Experiment" in src
    assert f"Experiment ID  : {exp.id}" in src
    assert "process-improve:" in src
    # Canonical import of the shared dispatch.
    assert "from process_improve.tool_spec import execute_tool_call" in src
    # Step comment and dispatch line.
    assert "# Step 1: generate_design" in src
    assert 'execute_tool_call("generate_design"' in src

    # Must be syntactically valid Python.
    compile(src, "<generated>", "exec")


async def test_script_roundtrip_reproduces_stored_tool_inputs(db_session, monkeypatch):
    stored_inputs = [
        {"design_type": "full_factorial", "n_factors": 2, "center_points": 1},
        {"model": "linear+interactions", "confidence_level": 0.95},
    ]
    exp = await _seed(
        db_session,
        tool_calls=[
            {"tool_name": "generate_design", "agent_turn": 1, "call_order": 1, "tool_input": stored_inputs[0]},
            {"tool_name": "evaluate_design", "agent_turn": 2, "call_order": 1, "tool_input": stored_inputs[1]},
        ],
    )
    payload = await reproducible_export_service.build_python_script(db_session, exp)
    src = payload.decode("utf-8")

    # Stub ``process_improve.tool_spec`` with a fake module that records
    # every (name, input) and returns a marker.  The generated script
    # imports from this module, so we install the stub in sys.modules
    # before exec'ing.
    recorded: list[tuple[str, dict]] = []

    def fake_execute_tool_call(name: str, tool_input: dict) -> dict:
        recorded.append((name, tool_input))
        return {"ok": True}

    fake_pi = types.ModuleType("process_improve")
    fake_ts = types.ModuleType("process_improve.tool_spec")
    fake_ts.execute_tool_call = fake_execute_tool_call
    fake_pi.tool_spec = fake_ts
    monkeypatch.setitem(sys.modules, "process_improve", fake_pi)
    monkeypatch.setitem(sys.modules, "process_improve.tool_spec", fake_ts)

    exec_globals: dict = {}
    exec(compile(src, "<roundtrip>", "exec"), exec_globals)  # noqa: S102

    assert [name for name, _ in recorded] == ["generate_design", "evaluate_design"]
    assert [payload_in for _, payload_in in recorded] == stored_inputs


async def test_triple_quote_in_experiment_name_is_sanitised(db_session):
    # A docstring-breaking name must not produce invalid Python.
    exp = await _seed(
        db_session,
        tool_calls=[
            {"tool_name": "generate_design", "agent_turn": 1, "call_order": 1, "tool_input": {}},
        ],
    )
    exp.name = 'Evil """ name'
    payload = await reproducible_export_service.build_python_script(db_session, exp)
    src = payload.decode("utf-8")
    # The raw triple-quote is replaced so the docstring stays single.
    assert "\"\"\"Reproducible analysis for experiment: Evil ''' name" in src
    compile(src, "<evil>", "exec")


# ---------------------------------------------------------------------------
# Notebook / literate-markdown / data-file / bundle tests (PR-2).
# ---------------------------------------------------------------------------


_DESIGN_INPUT_WITH_SEED = {
    "design_type": "full_factorial",
    "n_factors": 2,
    "seed": 42,
}
_EVAL_INPUT = {"model": "linear+interactions", "confidence_level": 0.95}


async def _seed_two_step_experiment(db_session):
    return await _seed(
        db_session,
        tool_calls=[
            {
                "tool_name": "generate_design",
                "agent_turn": 1,
                "call_order": 1,
                "tool_input": _DESIGN_INPUT_WITH_SEED,
            },
            {
                "tool_name": "evaluate_design",
                "agent_turn": 2,
                "call_order": 1,
                "tool_input": _EVAL_INPUT,
            },
        ],
    )


async def test_notebook_is_valid_nbformat(db_session):
    import nbformat  # noqa: PLC0415

    exp = await _seed_two_step_experiment(db_session)
    payload = await reproducible_export_service.build_notebook(db_session, exp)
    nb = nbformat.reads(payload.decode("utf-8"), as_version=4)
    # ``validate`` raises ``ValidationError`` on failure.
    nbformat.validate(nb)
    # Structure: header md + setup code + (md + code) per call + footer md.
    assert len(nb.cells) == 3 + 2 * 2
    assert nb.cells[0].cell_type == "markdown"
    assert nb.cells[1].cell_type == "code"
    assert "execute_tool_call" in nb.cells[1].source
    assert 'execute_tool_call("generate_design"' in nb.cells[3].source
    assert 'execute_tool_call("evaluate_design"' in nb.cells[5].source


async def test_notebook_raises_400_without_calls(db_session):
    exp = await _seed(db_session, tool_calls=[])
    with pytest.raises(HTTPException) as info:
        await reproducible_export_service.build_notebook(db_session, exp)
    assert info.value.status_code == 400


async def test_literate_markdown_contains_fenced_code_blocks(db_session):
    exp = await _seed_two_step_experiment(db_session)
    payload = await reproducible_export_service.build_literate_markdown(db_session, exp)
    text = payload.decode("utf-8")

    # Header metadata and setup block.
    assert "Reproducible analysis: My Experiment" in text
    assert 'pip install "process-improve==' in text
    # Both steps surface as their own fenced python blocks.
    assert text.count("```python") >= 3  # setup + 2 step blocks
    assert 'execute_tool_call("generate_design"' in text
    assert 'execute_tool_call("evaluate_design"' in text


async def test_literate_markdown_raises_400_without_calls(db_session):
    exp = await _seed(db_session, tool_calls=[])
    with pytest.raises(HTTPException) as info:
        await reproducible_export_service.build_literate_markdown(db_session, exp)
    assert info.value.status_code == 400


async def test_build_data_file_returns_valid_xlsx(db_session):
    from openpyxl import load_workbook  # noqa: PLC0415

    exp = await _seed_two_step_experiment(db_session)
    # Populate enough design/results data for the xlsx builder.
    exp.design_data = {
        "n_factors": 2,
        "n_runs": 2,
        "design_actual": [
            {"Temperature": 150, "Pressure": 1},
            {"Temperature": 200, "Pressure": 5},
        ],
    }
    exp.results_data = [{"run_index": 0, "yield": 85.2}]
    payload = reproducible_export_service.build_data_file(exp)
    wb = load_workbook(io.BytesIO(payload))
    assert "Design" in wb.sheetnames


async def test_build_requirements_txt_pins_version():
    assert reproducible_export_service.build_requirements_txt("1.5.1") == "process-improve==1.5.1\n"


async def test_collect_warnings_flags_missing_seed(db_session):
    exp = await _seed(
        db_session,
        tool_calls=[
            # No seed -> warning.
            {"tool_name": "generate_design", "agent_turn": 1, "call_order": 1, "tool_input": {"n_factors": 2}},
            # Seed present -> no warning.
            {
                "tool_name": "generate_design",
                "agent_turn": 2,
                "call_order": 1,
                "tool_input": {"n_factors": 3, "seed": 7},
            },
        ],
    )
    calls = await reproducible_export_service.fetch_analysis_tool_calls(db_session, exp)
    warnings = reproducible_export_service.collect_warnings(calls)
    assert len(warnings) == 1
    assert "no seed" in warnings[0]


async def test_bundle_contains_expected_entries(db_session):
    import zipfile  # noqa: PLC0415

    exp = await _seed_two_step_experiment(db_session)
    exp.design_data = {
        "n_factors": 2,
        "n_runs": 2,
        "design_actual": [
            {"Temperature": 150, "Pressure": 1},
            {"Temperature": 200, "Pressure": 5},
        ],
    }
    exp.results_data = [{"run_index": 0, "yield": 85.2}]
    payload = await reproducible_export_service.build_reproducible_bundle(db_session, exp)

    with zipfile.ZipFile(io.BytesIO(payload)) as zf:
        names = set(zf.namelist())
        assert names == {
            "analysis.py",
            "analysis.ipynb",
            "analysis.md",
            "data.xlsx",
            "README.md",
            "requirements.txt",
        }
        readme = zf.read("README.md").decode("utf-8")
        reqs = zf.read("requirements.txt").decode("utf-8")

    assert "Reproducible bundle" in readme
    assert reqs.startswith("process-improve==")


async def test_bundle_py_matches_standalone_py(db_session):
    import zipfile  # noqa: PLC0415

    exp = await _seed_two_step_experiment(db_session)
    # Force a stable pi_version / timestamp so both renders agree.
    pinned_version = "9.9.9"
    fixed_time = datetime(2026, 1, 1, tzinfo=UTC)

    def fake_resolve() -> str:
        return pinned_version

    class _FrozenDatetime(datetime):
        @classmethod
        def now(cls, tz=None):  # noqa: ARG003
            return fixed_time

    import app.services.reproducible_export_service as svc  # noqa: PLC0415

    orig_resolve = svc._resolve_process_improve_version
    orig_dt = svc.datetime
    svc._resolve_process_improve_version = fake_resolve
    svc.datetime = _FrozenDatetime
    try:
        standalone = await reproducible_export_service.build_python_script(db_session, exp)
        bundle = await reproducible_export_service.build_reproducible_bundle(db_session, exp)
    finally:
        svc._resolve_process_improve_version = orig_resolve
        svc.datetime = orig_dt

    with zipfile.ZipFile(io.BytesIO(bundle)) as zf:
        bundle_py = zf.read("analysis.py")

    assert bundle_py == standalone


async def test_bundle_readme_surfaces_seed_warning(db_session):
    import zipfile  # noqa: PLC0415

    exp = await _seed(
        db_session,
        tool_calls=[
            # generate_design without seed -> README must warn.
            {"tool_name": "generate_design", "agent_turn": 1, "call_order": 1, "tool_input": {"n_factors": 2}},
        ],
    )
    payload = await reproducible_export_service.build_reproducible_bundle(db_session, exp)
    with zipfile.ZipFile(io.BytesIO(payload)) as zf:
        readme = zf.read("README.md").decode("utf-8")
    assert "Warnings captured at export" in readme
    assert "no seed" in readme


async def test_bundle_raises_400_without_calls(db_session):
    exp = await _seed(db_session, tool_calls=[])
    with pytest.raises(HTTPException) as info:
        await reproducible_export_service.build_reproducible_bundle(db_session, exp)
    assert info.value.status_code == 400
