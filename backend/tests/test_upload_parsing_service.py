"""Tests for ``app.services.upload_parsing_service``."""

from __future__ import annotations

import datetime as dt
import io

import pytest
from openpyxl import Workbook

from app.services.upload_parsing_service import (
    UploadValidationError,
    parse_upload,
)


def _xlsx_bytes(rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# .xlsx
# ---------------------------------------------------------------------------


def test_parse_xlsx_standard_orientation() -> None:
    """Standard rows-as-experiments sheet round-trips to a 2D matrix."""

    raw = _xlsx_bytes(
        [
            ["Run", "Temperature", "Pressure", "Yield"],
            [1, 50, 1.0, 72.3],
            [2, 80, 1.0, 81.5],
            [3, 50, 2.0, 70.1],
            [4, 80, 2.0, 88.4],
        ]
    )

    result = parse_upload("design.xlsx", raw)

    assert result.detected_format == "xlsx"
    assert result.rows[0] == ["Run", "Temperature", "Pressure", "Yield"]
    assert len(result.rows) == 5
    assert result.rows[1][1] == 50


def test_parse_xlsx_transposed_orientation_kept_as_is() -> None:
    """Orientation detection is Claude's job — we just return the matrix."""

    raw = _xlsx_bytes(
        [
            ["Factor", "Run 1", "Run 2", "Run 3"],
            ["Temperature", 50, 80, 50],
            ["Pressure", 1.0, 1.0, 2.0],
        ]
    )

    result = parse_upload("transposed.xlsx", raw)

    assert result.rows[0] == ["Factor", "Run 1", "Run 2", "Run 3"]
    assert result.rows[1][0] == "Temperature"


def test_parse_xlsx_empty_design_header_only() -> None:
    """A header-only sheet (no data rows yet) parses to a single row."""

    raw = _xlsx_bytes([["Temperature", "Pressure", "Yield"]])

    result = parse_upload("empty.xlsx", raw)

    assert result.rows == [["Temperature", "Pressure", "Yield"]]


def test_parse_xlsx_coerces_datetime_to_isoformat() -> None:
    """openpyxl returns datetime objects; we serialise them for JSON."""

    raw = _xlsx_bytes(
        [
            ["Run", "Started"],
            [1, dt.datetime(2026, 4, 25, 10, 30)],
        ]
    )

    result = parse_upload("dates.xlsx", raw)

    assert result.rows[1][1] == "2026-04-25T10:30:00"


def test_parse_xlsx_pads_short_rows() -> None:
    """Different row widths are padded to the widest row."""

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "A"
    ws["B1"] = "B"
    ws["C1"] = "C"
    ws["A2"] = 1  # only one cell on this row
    buf = io.BytesIO()
    wb.save(buf)

    result = parse_upload("ragged.xlsx", buf.getvalue())

    assert all(len(row) == 3 for row in result.rows)


def test_parse_rejects_oversize_file() -> None:
    raw = b"\x00" * 1024
    with pytest.raises(UploadValidationError, match="bytes"):
        parse_upload("any.xlsx", raw, max_bytes=512, max_cells=10_000)


def test_parse_rejects_oversize_cell_count() -> None:
    raw = _xlsx_bytes([["A", "B", "C"], [1, 2, 3], [4, 5, 6]])
    with pytest.raises(UploadValidationError, match="cells"):
        parse_upload("small.xlsx", raw, max_cells=5)


def test_parse_rejects_legacy_xls() -> None:
    with pytest.raises(UploadValidationError, match="\\.xlsx"):
        parse_upload("old.xls", b"irrelevant")


def test_parse_rejects_unknown_extension() -> None:
    with pytest.raises(UploadValidationError, match="extension"):
        parse_upload("design.txt", b"a,b,c\n1,2,3\n")


def test_parse_rejects_corrupt_xlsx() -> None:
    with pytest.raises(UploadValidationError, match="Excel"):
        parse_upload("bad.xlsx", b"not actually xlsx")


# ---------------------------------------------------------------------------
# .csv
# ---------------------------------------------------------------------------


def test_parse_csv_utf8() -> None:
    raw = b"Temperature,Pressure,Yield\n50,1.0,72.3\n80,2.0,88.4\n"
    result = parse_upload("design.csv", raw)
    assert result.detected_format == "csv"
    assert result.rows[0] == ["Temperature", "Pressure", "Yield"]
    assert result.rows[1] == ["50", "1.0", "72.3"]


def test_parse_csv_utf8_with_bom() -> None:
    raw = b"\xef\xbb\xbfA,B\n1,2\n"
    result = parse_upload("bom.csv", raw)
    # First cell would be "﻿A" without utf-8-sig handling.
    assert result.rows[0][0] == "A"


def test_parse_csv_latin1_fallback() -> None:
    """A latin-1-encoded file with non-ASCII still parses without exploding."""

    raw = "Naïve,Yield\n50,72\n".encode("latin-1")
    result = parse_upload("latin1.csv", raw)
    assert result.rows[0] == ["Naïve", "Yield"]


def test_parse_csv_ragged_rows_are_padded() -> None:
    raw = b"A,B,C\n1,2\n3,4,5\n"
    result = parse_upload("ragged.csv", raw)
    assert all(len(row) == 3 for row in result.rows)


def test_parse_csv_drops_trailing_blank_rows() -> None:
    raw = b"A,B\n1,2\n,\n,\n"
    result = parse_upload("trailing.csv", raw)
    assert result.rows == [["A", "B"], ["1", "2"]]
