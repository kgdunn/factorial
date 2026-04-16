"""Unit tests for ``app.services.export_service``.

Each test calls the ``build_*`` functions directly with a fake
Experiment object and asserts the output shape.  No Playwright or
WeasyPrint is invoked by these tests — those pieces are mocked or
asserted to be gated by ``settings.exports_enable_pdf``.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, datetime
from typing import Any
from unittest.mock import patch

import pytest

from app.services import export_service


class _FakeExperiment:
    def __init__(self, **overrides: Any) -> None:
        self.id = overrides.get("id", uuid.uuid4())
        self.name = overrides.get("name", "My Experiment")
        self.status = overrides.get("status", "active")
        self.design_type = overrides.get("design_type", "full_factorial")
        self.factors = overrides.get(
            "factors",
            [
                {"name": "Temperature", "low": 150, "high": 200},
                {"name": "Pressure", "low": 1, "high": 5},
            ],
        )
        self.design_data = overrides.get(
            "design_data",
            {
                "design_type": "full_factorial",
                "n_factors": 2,
                "n_runs": 4,
                "factor_names": ["Temperature", "Pressure"],
                "design_actual": [
                    {"Temperature": 150, "Pressure": 1},
                    {"Temperature": 200, "Pressure": 1},
                    {"Temperature": 150, "Pressure": 5},
                    {"Temperature": 200, "Pressure": 5},
                ],
            },
        )
        self.results_data = overrides.get(
            "results_data",
            [
                {"run_index": 0, "yield": 85.2},
                {"run_index": 1, "yield": 91.0},
            ],
        )
        self.created_at = overrides.get("created_at", datetime.now(UTC))
        self.user_id = overrides.get("user_id")


class TestCsvExport:
    def test_csv_round_trips(self):
        exp = _FakeExperiment()
        payload = export_service.build_csv(exp)
        reader = csv.DictReader(io.StringIO(payload.decode("utf-8")))
        rows = list(reader)
        assert len(rows) == 4
        assert "Temperature" in rows[0]
        assert "Pressure" in rows[0]
        assert "yield" in rows[0]
        assert rows[0]["yield"] == "85.2"
        assert rows[2]["yield"] == ""  # no results entered for run 2

    def test_csv_omits_results_when_disabled(self):
        exp = _FakeExperiment()
        payload = export_service.build_csv(exp, include_results=False)
        reader = csv.DictReader(io.StringIO(payload.decode("utf-8")))
        assert "yield" not in (reader.fieldnames or [])


class TestXlsxExport:
    def test_xlsx_opens_and_has_tabs(self):
        from openpyxl import load_workbook  # noqa: PLC0415

        exp = _FakeExperiment()
        payload = export_service.build_xlsx(exp)
        wb = load_workbook(io.BytesIO(payload))
        assert "Design" in wb.sheetnames
        assert "Responses" in wb.sheetnames
        assert "Metadata" in wb.sheetnames
        design = wb["Design"]
        header = [c.value for c in design[1]]
        assert "Temperature" in header

    def test_xlsx_skips_responses_tab_when_disabled(self):
        from openpyxl import load_workbook  # noqa: PLC0415

        exp = _FakeExperiment()
        payload = export_service.build_xlsx(exp, include_results=False)
        wb = load_workbook(io.BytesIO(payload))
        assert "Design" in wb.sheetnames
        assert "Responses" not in wb.sheetnames


class TestMarkdownExport:
    def test_markdown_contains_factor_names(self):
        exp = _FakeExperiment()
        payload = export_service.build_markdown(exp, share_url="https://example.com/share/tok")
        text = payload.decode("utf-8")
        assert "My Experiment" in text
        assert "Temperature" in text
        assert "Pressure" in text
        assert "https://example.com/share/tok" in text


class TestPdfExport:
    @pytest.mark.asyncio
    async def test_pdf_disabled_returns_503(self):
        from fastapi import HTTPException  # noqa: PLC0415

        exp = _FakeExperiment()
        with patch.object(export_service, "settings") as mock_settings:
            mock_settings.exports_enable_pdf = False
            with pytest.raises(HTTPException) as info:
                await export_service.build_pdf(exp)
        assert info.value.status_code == 503

    @pytest.mark.asyncio
    async def test_pdf_missing_weasyprint_returns_503(self):
        from fastapi import HTTPException  # noqa: PLC0415

        exp = _FakeExperiment()
        with (
            patch.object(export_service, "settings") as mock_settings,
            patch.dict("sys.modules", {"weasyprint": None}),
        ):
            mock_settings.exports_enable_pdf = True
            with pytest.raises(HTTPException) as info:
                await export_service.build_pdf(exp)
        assert info.value.status_code == 503
